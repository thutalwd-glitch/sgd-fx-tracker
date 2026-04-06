# =============================================================================
# SGD Exchange Rate Analyser
# Author: Thuta Lin
# Description: Downloads 3 years of SGD exchange rate data, analyses trends,
#              and exports a formatted Excel report.
# =============================================================================

# --- STEP 1: IMPORT LIBRARIES ---
# Think of libraries like toolboxes. Each one gives us extra abilities.

import matplotlib
matplotlib.use("Agg")  # IMPORTANT: Use non-interactive backend — required when running
                       # in the cloud (GitHub Actions) where there is no screen/display.
                       # Without this line, matplotlib crashes in headless environments.

import yfinance as yf          # Downloads financial data from Yahoo Finance (free)
yf.set_tz_cache_location("/tmp")  # Fix for GitHub Actions: redirect yfinance's timezone
                                   # cache to /tmp so it doesn't hit a "database is locked"
                                   # error when running in the cloud.
import pandas as pd            # The main tool for working with tables of data
import matplotlib.pyplot as plt  # Used to draw charts/graphs
import matplotlib.dates as mdates  # Helps format dates on chart axes
from openpyxl import Workbook  # Creates and edits Excel files
from openpyxl.styles import (  # These let us format cells (bold, colour, etc.)
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter  # Converts column numbers to letters (e.g. 1 → "A")
from openpyxl.chart import LineChart, Reference  # For embedding charts inside Excel
import os                      # Helps us work with file paths and folders
from datetime import datetime  # Used to get today's date

# --- STEP 2: SETTINGS ---
# These are the variables we might want to change in the future.
# Keeping them here at the top makes the code easy to update.

# The currency pairs we want to track.
# In finance, "SGDUSD=X" means: how many USD does 1 SGD buy?
TICKERS = {
    "SGDUSD=X": "SGD/USD",
    "SGDEUR=X": "SGD/EUR",
    "SGDGBP=X": "SGD/GBP",
    "SGDJPY=X": "SGD/JPY",
    "SGDAUD=X": "SGD/AUD",
    "SGDCNY=X": "SGD/CNY",
}

PERIOD = "3y"          # How far back to fetch data (3 years)
OUTPUT_FOLDER = "output"   # Folder where we save the Excel file


# =============================================================================
# SECTION 1: DOWNLOAD DATA
# =============================================================================

def download_data(tickers, period):
    """
    Downloads historical exchange rate data from Yahoo Finance.

    What this does in plain English:
    - We give it a list of currency codes (like "SGDUSD=X")
    - It goes to Yahoo Finance and downloads daily closing prices
    - It returns a clean table (called a DataFrame) with dates as rows
      and currency pairs as columns
    """
    print("📥 Downloading exchange rate data from Yahoo Finance...")

    # yf.download fetches the data. We only want the 'Close' price (end-of-day rate)
    raw = yf.download(list(tickers.keys()), period=period, auto_adjust=True, progress=False)
    df = raw["Close"]  # "Close" = the exchange rate at the end of each trading day

    # Rename columns from ticker codes to friendly names (e.g. "SGDUSD=X" → "SGD/USD")
    df = df.rename(columns=tickers)

    # Drop any rows where ALL values are missing (weekends/public holidays have no data)
    df = df.dropna(how="all")

    # Make sure the date column is formatted as a proper date (not a string)
    df.index = pd.to_datetime(df.index)

    print(f"✅ Downloaded {len(df)} days of data ({df.index[0].date()} to {df.index[-1].date()})")
    return df


# =============================================================================
# SECTION 2: ANALYSE THE DATA
# =============================================================================

def analyse_data(df):
    """
    Runs analysis on the exchange rate data.

    Returns a dictionary with:
    - summary: key stats (mean, min, max, volatility, etc.) for each currency
    - monthly: average exchange rate per month for each currency
    - rolling: 30-day rolling average (smooths out daily noise to show trends)
    - pct_change: daily % change (tells us how much the rate moved each day)
    """
    print("🔍 Analysing data...")

    results = {}

    # --- Summary Statistics ---
    # For each currency, calculate important numbers that describe the data
    summary_rows = []
    for col in df.columns:
        series = df[col].dropna()  # Remove any empty rows for this currency

        # Calculate year-over-year change:
        # Compare today's rate to the rate exactly 1 year ago
        if len(series) > 252:  # 252 = approximate trading days in a year
            rate_now = series.iloc[-1]       # Most recent rate
            rate_1y_ago = series.iloc[-252]  # Rate ~1 year ago
            yoy_change = ((rate_now - rate_1y_ago) / rate_1y_ago) * 100  # % change
        else:
            yoy_change = None

        summary_rows.append({
            "Currency Pair": col,
            "Latest Rate": round(series.iloc[-1], 4),
            "3Y Average": round(series.mean(), 4),
            "3Y High": round(series.max(), 4),
            "3Y Low": round(series.min(), 4),
            # Volatility = standard deviation = how much the rate jumps around daily
            # Higher number = more unstable currency pair
            "Volatility (Std Dev)": round(series.std(), 4),
            "YoY Change (%)": round(yoy_change, 2) if yoy_change is not None else "N/A",
        })

    results["summary"] = pd.DataFrame(summary_rows)

    # --- Monthly Averages ---
    # Group by year-month and calculate average rate for each month
    # This smooths out day-to-day noise so we can see longer trends
    monthly = df.resample("ME").mean()  # "ME" = Month End
    monthly.index = monthly.index.strftime("%b %Y")  # Format as "Jan 2023"
    results["monthly"] = monthly

    # --- 30-Day Rolling Average ---
    # Instead of showing every single day's rate (very noisy),
    # we average the last 30 days. This shows the underlying trend more clearly.
    results["rolling"] = df.rolling(window=30).mean()

    # --- Indexed / Normalised Data (set to 100 at start) ---
    # This is a standard finance technique called "indexing".
    # We take the first valid value for each currency and divide every row by it,
    # then multiply by 100. So every currency STARTS at 100.
    # After that, if SGD/USD reads 105, it means USD has moved up 5% from the start.
    # If SGD/JPY reads 112, it means JPY moved up 12%.
    # This lets us compare currencies on the SAME scale — even though JPY is ~100x bigger.
    first_valid = df.apply(lambda col: col.dropna().iloc[0])  # First value per currency
    indexed = (df / first_valid) * 100                         # Normalise to 100
    results["indexed"] = indexed.rolling(window=30).mean()     # Smooth with 30-day average

    # --- Daily % Change ---
    # How much did the rate move compared to yesterday? (as a percentage)
    results["pct_change"] = df.pct_change() * 100

    print("✅ Analysis complete")
    return results


# =============================================================================
# SECTION 3: CREATE CHARTS
# =============================================================================

def create_charts(df, results, output_folder):
    """
    Creates and saves 2 chart images as PNG files.
    These will also be embedded into the Excel report later.

    Chart 1: Line chart showing all 6 exchange rates over 3 years
    Chart 2: Bar chart showing year-over-year % change for each currency
    """
    print("📊 Creating charts...")

    chart_paths = []

    # ---- CHART 1: Indexed Exchange Rate Trends (Normalised to 100) ----
    #
    # Instead of plotting raw rates (which makes JPY dominate at ~100x scale),
    # we index every currency to 100 at the start date.
    # This shows RELATIVE movement — which currency moved most, and in which direction.
    #
    # Reading the chart:
    #   - All lines start at 100
    #   - Above 100 = SGD has strengthened against that currency since the start
    #   - Below 100 = SGD has weakened against that currency since the start
    #   - A line at 112 means a +12% move; a line at 95 means a -5% move

    indexed = results["indexed"]  # 30-day smoothed, normalised data

    fig, ax = plt.subplots(figsize=(14, 6))

    # Use a distinct colour palette so every line is clearly different
    colors = ["#2980b9", "#e67e22", "#27ae60", "#8e44ad", "#e74c3c", "#16a085"]

    for col, color in zip(indexed.columns, colors):
        ax.plot(indexed.index, indexed[col], label=col, linewidth=2, color=color)

    # Horizontal reference line at 100 (= no change from start)
    ax.axhline(100, color="black", linewidth=1, linestyle="--", alpha=0.5, label="Baseline (100)")

    # Shade the area above/below 100 very lightly to make it easier to read
    ax.fill_between(indexed.index, 100, indexed.max(axis=1), alpha=0.03, color="green")
    ax.fill_between(indexed.index, indexed.min(axis=1), 100, alpha=0.03, color="red")

    ax.set_title("SGD Exchange Rate Trends — Indexed to 100 (3-Year, 30-Day Rolling Average)",
                 fontsize=13, fontweight="bold", pad=15)
    ax.set_xlabel("Date", fontsize=11)
    ax.set_ylabel("Index (Start = 100)", fontsize=11)

    # Add annotation explaining what the index means
    ax.text(0.01, 0.97,
            "Above 100 = SGD strengthened  |  Below 100 = SGD weakened  |  All currencies start at same scale",
            transform=ax.transAxes, fontsize=8, va="top", color="grey")

    ax.legend(loc="upper left", fontsize=9, framealpha=0.9)
    ax.grid(True, alpha=0.3)

    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    ax.xaxis.set_major_locator(mdates.YearLocator())
    fig.autofmt_xdate()

    plt.tight_layout()
    chart1_path = os.path.join(output_folder, "chart1_trends.png")
    plt.savefig(chart1_path, dpi=150, bbox_inches="tight")
    plt.close()
    chart_paths.append(chart1_path)
    print(f"  ✅ Chart 1 saved: {chart1_path}")

    # ---- CHART 2: Year-over-Year Change (Bar Chart) — Improved ----
    summary = results["summary"]

    # Filter rows where YoY change exists and convert to float
    yoy_data = summary[summary["YoY Change (%)"] != "N/A"].copy()
    yoy_data["YoY Change (%)"] = yoy_data["YoY Change (%)"].astype(float)

    # Sort bars from most negative → most positive (left to right tells a clearer story)
    yoy_data = yoy_data.sort_values("YoY Change (%)")

    # Colour palette: muted tones look more professional than bright green/red
    POS_COLOR = "#27ae60"   # Confident green
    NEG_COLOR = "#c0392b"   # Firm red
    colors = [POS_COLOR if v >= 0 else NEG_COLOR for v in yoy_data["YoY Change (%)"]]

    fig, ax = plt.subplots(figsize=(11, 5.5))

    # Light grey plot background — modern dashboard look
    ax.set_facecolor("#f7f9fb")
    fig.patch.set_facecolor("#ffffff")

    bars = ax.bar(
        yoy_data["Currency Pair"],
        yoy_data["YoY Change (%)"],
        color=colors,
        edgecolor="white",
        linewidth=1.2,
        width=0.55,
        zorder=3   # Draw bars on top of grid lines
    )

    # --- Value labels: always placed OUTSIDE the bar, never overlapping ---
    # For positive bars: label sits above the bar top
    # For negative bars: label sits below the bar bottom
    # We add a small padding (offset) so the label never touches the bar
    y_range = yoy_data["YoY Change (%)"].max() - yoy_data["YoY Change (%)"].min()
    offset = y_range * 0.03  # 3% of the total range = consistent padding

    for bar, val in zip(bars, yoy_data["YoY Change (%)"]):
        x_center = bar.get_x() + bar.get_width() / 2
        if val >= 0:
            y_pos = bar.get_height() + offset
            va = "bottom"
        else:
            y_pos = bar.get_height() - offset
            va = "top"
        ax.text(
            x_center, y_pos,
            f"{val:+.2f}%",
            ha="center", va=va,
            fontsize=10, fontweight="bold",
            color=POS_COLOR if val >= 0 else NEG_COLOR
        )

    # Zero baseline — slightly thicker and dark so it's clearly the reference point
    ax.axhline(0, color="#2c3e50", linewidth=1.2, zorder=4)

    # Grid lines behind bars (zorder=2 puts them under bars)
    ax.yaxis.grid(True, color="#dce3ea", linewidth=0.8, zorder=2)
    ax.set_axisbelow(True)

    # Remove the top and right spines (chart borders) — cleaner look
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#cccccc")
    ax.spines["bottom"].set_color("#cccccc")

    # Titles and labels
    ax.set_title("SGD Year-over-Year Performance vs Major Currencies",
                 fontsize=14, fontweight="bold", color="#1a1a2e", pad=18)
    ax.set_xlabel("Currency Pair", fontsize=11, color="#444444", labelpad=10)
    ax.set_ylabel("Change (%)", fontsize=11, color="#444444", labelpad=10)
    ax.tick_params(axis="x", labelsize=10, colors="#333333")
    ax.tick_params(axis="y", labelsize=9, colors="#666666")

    # Legend note in the bottom-right corner — cleaner placement than top-left
    ax.text(0.99, 0.02,
            "▲ Green = SGD strengthened   ▼ Red = SGD weakened",
            transform=ax.transAxes, fontsize=8, va="bottom", ha="right",
            color="#888888", style="italic")

    plt.tight_layout(pad=1.5)
    chart2_path = os.path.join(output_folder, "chart2_yoy.png")
    plt.savefig(chart2_path, dpi=150, bbox_inches="tight")
    plt.close()
    chart_paths.append(chart2_path)
    print(f"  ✅ Chart 2 saved: {chart2_path}")

    return chart_paths


# =============================================================================
# SECTION 4: EXPORT TO EXCEL
# =============================================================================

def export_to_excel(df, results, chart_paths, output_folder):
    """
    Creates a formatted Excel workbook with 4 sheets:
    - Sheet 1: Summary Statistics  (key numbers for each currency)
    - Sheet 2: Historical Data      (all daily rates for 3 years)
    - Sheet 3: Monthly Averages     (rates averaged by month)
    - Sheet 4: Charts               (the 2 charts embedded as images)
    """
    print("📁 Exporting to Excel...")

    # --- Helper: Define consistent colour/style theme ---
    HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # Dark navy header
    ACCENT_FILL  = PatternFill("solid", fgColor="D6E4F0")   # Light blue for alternate rows
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)  # White bold text for headers
    NORMAL_FONT  = Font(size=10)
    CENTER       = Alignment(horizontal="center", vertical="center")
    LEFT         = Alignment(horizontal="left", vertical="center")
    THIN_BORDER  = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def style_header_row(ws, row_num, num_cols):
        """Applies navy background + white bold text to the header row."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    def style_data_rows(ws, start_row, end_row, num_cols):
        """Applies light blue alternating rows and border to data cells."""
        for row in range(start_row, end_row + 1):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                if row % 2 == 0:  # Every other row gets a light blue background
                    cell.fill = ACCENT_FILL
                cell.font = NORMAL_FONT
                cell.alignment = CENTER
                cell.border = THIN_BORDER

    def set_col_widths(ws, widths):
        """Sets column widths. widths is a dict like {1: 20, 2: 15}"""
        for col_num, width in widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width

    # Create the workbook (Excel file)
    wb = Workbook()

    # -----------------------------------------------------------------------
    # SHEET 1: Summary Statistics
    # -----------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Summary"

    # Title row
    ws1.merge_cells("A1:G1")  # Merge cells A1 to G1 to make a wide title
    title_cell = ws1["A1"]
    title_cell.value = f"SGD Exchange Rate — Summary Statistics  |  Generated: {datetime.now().strftime('%d %b %Y')}"
    title_cell.font = Font(bold=True, size=13, color="1F3864")
    title_cell.alignment = LEFT
    ws1.row_dimensions[1].height = 30

    # Blank row for spacing
    ws1.row_dimensions[2].height = 8

    # Write column headers in row 3
    summary_df = results["summary"]
    headers = list(summary_df.columns)
    for col_i, header in enumerate(headers, start=1):
        ws1.cell(row=3, column=col_i, value=header)
    style_header_row(ws1, 3, len(headers))

    # Write the data rows (starting row 4)
    for row_i, row_data in enumerate(summary_df.itertuples(index=False), start=4):
        for col_i, value in enumerate(row_data, start=1):
            ws1.cell(row=row_i, column=col_i, value=value)
    style_data_rows(ws1, 4, 3 + len(summary_df), len(headers))

    # Set column widths for readability
    set_col_widths(ws1, {1: 18, 2: 14, 3: 14, 4: 12, 5: 12, 6: 20, 7: 18})

    # -----------------------------------------------------------------------
    # SHEET 2: Historical Data (all daily rates)
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("Historical Data")

    # We'll write dates + one column per currency
    ws2.cell(row=1, column=1, value="Date")
    for col_i, col_name in enumerate(df.columns, start=2):
        ws2.cell(row=1, column=col_i, value=col_name)
    style_header_row(ws2, 1, len(df.columns) + 1)

    # Write each day's data
    for row_i, (date, row_vals) in enumerate(df.iterrows(), start=2):
        ws2.cell(row=row_i, column=1, value=date.strftime("%Y-%m-%d"))
        for col_i, val in enumerate(row_vals, start=2):
            ws2.cell(row=row_i, column=col_i, value=round(float(val), 4) if pd.notna(val) else "")

    # Only style the first 50 visible rows to keep the file size small
    style_data_rows(ws2, 2, min(len(df) + 1, 51), len(df.columns) + 1)
    set_col_widths(ws2, {i: 14 for i in range(1, len(df.columns) + 2)})

    # -----------------------------------------------------------------------
    # SHEET 3: Monthly Averages
    # -----------------------------------------------------------------------
    ws3 = wb.create_sheet("Monthly Averages")

    monthly_df = results["monthly"]

    ws3.cell(row=1, column=1, value="Month")
    for col_i, col_name in enumerate(monthly_df.columns, start=2):
        ws3.cell(row=1, column=col_i, value=col_name)
    style_header_row(ws3, 1, len(monthly_df.columns) + 1)

    for row_i, (month, row_vals) in enumerate(monthly_df.iterrows(), start=2):
        ws3.cell(row=row_i, column=1, value=month)
        for col_i, val in enumerate(row_vals, start=2):
            ws3.cell(row=row_i, column=col_i, value=round(float(val), 4) if pd.notna(val) else "")

    style_data_rows(ws3, 2, len(monthly_df) + 1, len(monthly_df.columns) + 1)
    set_col_widths(ws3, {i: 14 for i in range(1, len(monthly_df.columns) + 2)})

    # -----------------------------------------------------------------------
    # SHEET 4: Charts
    # -----------------------------------------------------------------------
    ws4 = wb.create_sheet("Charts")

    ws4.merge_cells("A1:N1")
    ws4["A1"].value = "SGD Exchange Rate — Visual Analysis"
    ws4["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws4["A1"].alignment = LEFT
    ws4.row_dimensions[1].height = 30

    # Embed the PNG charts into the Excel sheet
    # openpyxl lets us place images at specific cell coordinates
    from openpyxl.drawing.image import Image as XLImage

    if len(chart_paths) > 0 and os.path.exists(chart_paths[0]):
        img1 = XLImage(chart_paths[0])
        img1.width = 750
        img1.height = 320
        ws4.add_image(img1, "A3")   # Place chart 1 at cell A3

    if len(chart_paths) > 1 and os.path.exists(chart_paths[1]):
        img2 = XLImage(chart_paths[1])
        img2.width = 560
        img2.height = 280
        ws4.add_image(img2, "A22")  # Place chart 2 below chart 1

    # Save the workbook to the output folder
    filename = f"SGD_FX_Analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_folder, filename)
    wb.save(filepath)

    print(f"✅ Excel report saved: {filepath}")
    return filepath


# =============================================================================
# MAIN — This is where everything runs
# =============================================================================

if __name__ == "__main__":
    # "__main__" means: only run this block if you run THIS file directly.
    # (If someone imports this file as a library, this block won't run.)

    print("=" * 60)
    print("  SGD Exchange Rate Analyser")
    print("=" * 60)

    # Make sure the output folder exists (create it if not)
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    # Run each step in order
    df = download_data(TICKERS, PERIOD)           # Step 1: Get data
    results = analyse_data(df)                    # Step 2: Analyse
    chart_paths = create_charts(df, results, OUTPUT_FOLDER)   # Step 3: Charts
    filepath = export_to_excel(df, results, chart_paths, OUTPUT_FOLDER)  # Step 4: Excel

    print()
    print("=" * 60)
    print(f"  ✅ Done! Report saved to: {filepath}")
    print("=" * 60)
